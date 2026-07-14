"""Integration tests for Excel, PDF, and chart generation."""

import json
from pathlib import Path

import pytest

from charts.generator import ChartGenerator
from excel_generator.workbook import ExcelReportGenerator
from financial_engine.engine import FinancialPlanningEngine
from financial_engine.models import ClientFinancialProfile
from pdf_generator.report import PDFReportGenerator

ROOT = Path(__file__).resolve().parents[1]
FIRM = {
    "name": "WealthCraft Advisors",
    "tagline": "Clarity. Discipline. Prosperity.",
    "email": "advisor@wealthcraft.example",
    "phone": "+91-00000-00000",
}


@pytest.fixture(scope="module")
def plan_and_profile():
    engine = FinancialPlanningEngine(config_dir=ROOT / "config")
    data = json.loads((ROOT / "sample_data" / "demo_client.json").read_text())
    profile = ClientFinancialProfile.from_dict(data)
    plan = engine.plan_to_dict(engine.generate_plan(profile))
    return plan, profile


def test_charts(tmp_path, plan_and_profile):
    plan, _ = plan_and_profile
    paths = ChartGenerator(tmp_path / "charts").generate_all(plan)
    assert len(paths) >= 5
    for p in paths:
        assert Path(p).exists()
        assert Path(p).stat().st_size > 1000


def test_excel_workbook(tmp_path, plan_and_profile):
    plan, profile = plan_and_profile
    out = tmp_path / "plan.xlsx"
    ExcelReportGenerator(FIRM).generate(plan, profile, out)
    assert out.exists()
    assert out.stat().st_size > 20_000

    from openpyxl import load_workbook

    wb = load_workbook(out)
    required = {
        "Cover",
        "Client Summary",
        "Dashboard",
        "Goal Planning",
        "Retirement",
        "Tax Planning",
        "Recommendations",
        "Glossary",
        "Assumptions",
    }
    assert required.issubset(set(wb.sheetnames))


def test_pdf_report(tmp_path, plan_and_profile):
    plan, profile = plan_and_profile
    charts = ChartGenerator(tmp_path / "charts").generate_all(plan)
    out = tmp_path / "plan.pdf"
    PDFReportGenerator(FIRM).generate(plan, profile, out, charts)
    assert out.exists()
    assert out.stat().st_size > 30_000
    raw = out.read_bytes()
    assert b"DejaVu" in raw or b"DejaVuSans" in raw


def test_pdf_rupee_symbol_renders(tmp_path, plan_and_profile):
    """Helvetica lacks ₹; embedded DejaVu must make the symbol extractable."""
    pytest.importorskip("pypdf")
    from pypdf import PdfReader

    plan, profile = plan_and_profile
    out = tmp_path / "rupee.pdf"
    PDFReportGenerator(FIRM).generate(plan, profile, out, [])
    text = "\n".join((p.extract_text() or "") for p in PdfReader(str(out)).pages[:5])
    assert "₹" in text
    assert text.count("₹") >= 3
