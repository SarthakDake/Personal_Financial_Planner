"""
Professional multi-sheet Excel financial plan workbook.

Sheets mirror consulting-firm deliverables with formatting, charts,
freeze panes, filters, and named ranges.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from excel_generator.styles import (
    ACCENT_FILL,
    ALT_FILL,
    BAD_FILL,
    GOOD_FILL,
    HEADER_FILL,
    LIGHT_FILL,
    SECTION_FONT,
    THIN_BORDER,
    TITLE_FONT,
    VALUE_FONT,
    WARN_FILL,
    apply_header_row,
    autosize,
    write_kv,
)
from financial_engine.models import ClientFinancialProfile
from utils.money import format_inr


class ExcelReportGenerator:
    """Generates a comprehensive wealth-management Excel workbook."""

    def __init__(self, firm: dict[str, str]):
        self.firm = firm

    def generate(
        self,
        plan: dict[str, Any],
        profile: ClientFinancialProfile,
        output_path: str | Path,
        chart_paths: list[str] | None = None,
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Remove default sheet after creating cover
        default = wb.active
        default.title = "Cover"

        self._cover(wb["Cover"], plan, profile)
        self._client_summary(wb.create_sheet("Client Summary"), plan, profile)
        self._snapshot(wb.create_sheet("Financial Snapshot"), plan)
        self._income(wb.create_sheet("Income"), plan, profile)
        self._expenses(wb.create_sheet("Expenses"), plan, profile)
        self._assets(wb.create_sheet("Assets"), plan, profile)
        self._liabilities(wb.create_sheet("Liabilities"), plan)
        self._net_worth(wb.create_sheet("Net Worth"), plan)
        self._cash_flow(wb.create_sheet("Cash Flow"), plan)
        self._emergency(wb.create_sheet("Emergency Fund"), plan)
        self._insurance(wb.create_sheet("Insurance"), plan)
        self._portfolio(wb.create_sheet("Investment Portfolio"), plan)
        self._allocation(wb.create_sheet("Asset Allocation"), plan)
        self._mutual_funds(wb.create_sheet("Mutual Funds"), plan, profile)
        self._stocks(wb.create_sheet("Stocks"), plan, profile)
        self._gold(wb.create_sheet("Gold"), plan, profile)
        self._loans(wb.create_sheet("Loans"), plan)
        self._loan_schedule(wb.create_sheet("Loan Schedule"), plan)
        self._goals(wb.create_sheet("Goal Planning"), plan)
        self._retirement(wb.create_sheet("Retirement"), plan)
        self._tax(wb.create_sheet("Tax Planning"), plan)
        self._education(wb.create_sheet("Education Planning"), plan)
        self._marriage(wb.create_sheet("Marriage Planning"), plan)
        self._travel(wb.create_sheet("Travel Planning"), plan)
        self._estate(wb.create_sheet("Estate Planning"), plan)
        self._recommendations(wb.create_sheet("Recommendations"), plan)
        self._risk(wb.create_sheet("Risk Analysis"), plan)
        self._scenarios(wb.create_sheet("Scenario Analysis"), plan)
        self._inflation(wb.create_sheet("Inflation Analysis"), plan)
        self._dashboard(wb.create_sheet("Dashboard"), plan)
        self._charts_sheet(wb.create_sheet("Charts"), plan)
        self._assumptions(wb.create_sheet("Assumptions"), plan)
        self._glossary(wb.create_sheet("Glossary"))

        # Print setup for all sheets
        for ws in wb.worksheets:
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = "1:1"
            ws.freeze_panes = "A2"

        wb["Cover"].freeze_panes = None
        wb["Dashboard"].freeze_panes = "A2"

        wb.save(output_path)
        return output_path

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _cover(self, ws, plan, profile):
        ws.merge_cells("B2:G2")
        ws["B2"] = self.firm.get("name", "WealthCraft Advisors")
        ws["B2"].font = TITLE_FONT
        ws.merge_cells("B3:G3")
        ws["B3"] = self.firm.get("tagline", "")
        ws["B3"].font = Font(name="Calibri", size=12, italic=True, color="1F6F8B")

        ws.merge_cells("B5:G5")
        ws["B5"] = "COMPREHENSIVE FINANCIAL PLAN"
        ws["B5"].font = Font(name="Calibri", size=18, bold=True, color="0B3D5C")

        row = 8
        row = write_kv(ws, row, "Client Name", profile.personal.full_name or "—")
        row = write_kv(ws, row, "Age", profile.personal.age)
        row = write_kv(ws, row, "Retirement Age", profile.personal.retirement_age)
        row = write_kv(ws, row, "Risk Profile", plan["summary"].get("risk_profile", "").title())
        row = write_kv(ws, row, "Plan Date", datetime.now().strftime("%d %B %Y"))
        row = write_kv(ws, row, "Health Score", f"{plan['health_score']['score']} ({plan['health_score']['grade']})")
        row = write_kv(ws, row, "Net Worth", format_inr(plan["net_worth"]["net_worth"]))
        row += 2
        row = write_kv(ws, row, "Advisor Contact", self.firm.get("email", ""))
        row = write_kv(ws, row, "Phone", self.firm.get("phone", ""))
        ws.merge_cells("B20:G22")
        ws["B20"] = (
            "Confidential — This document is prepared for the named client only. "
            "Projections are based on configurable assumptions and are not guarantees of future performance. "
            "Please consult your SEBI-registered advisor before acting."
        )
        autosize(ws)

    def _client_summary(self, ws, plan, profile):
        ws["A1"] = "Client Summary"
        ws["A1"].font = TITLE_FONT
        p = profile.personal
        fields = [
            ("Full Name", p.full_name),
            ("Age", p.age),
            ("Retirement Age", p.retirement_age),
            ("Life Expectancy", p.life_expectancy),
            ("Marital Status", p.marital_status.value),
            ("Dependents", p.dependents),
            ("Children", p.children),
            ("City Type", p.city_type),
            ("Email", p.email),
            ("Phone", p.phone),
            ("PAN", p.pan),
            ("Risk Profile", plan["risk"].get("profile", "")),
            ("Investment Personality", plan["risk"].get("personality", "")),
        ]
        row = 3
        apply_header_row(ws, 3, 1, 2)
        ws["A3"], ws["B3"] = "Field", "Value"
        row = 4
        for label, value in fields:
            ws.cell(row=row, column=1, value=label).border = THIN_BORDER
            ws.cell(row=row, column=2, value=value).border = THIN_BORDER
            if row % 2 == 0:
                ws.cell(row=row, column=1).fill = ALT_FILL
                ws.cell(row=row, column=2).fill = ALT_FILL
            row += 1
        autosize(ws)
        ws.auto_filter.ref = f"A3:B{row - 1}"

    def _snapshot(self, ws, plan):
        ws["A1"] = "Financial Snapshot"
        ws["A1"].font = TITLE_FONT
        s = plan["summary"]
        items = [
            ("Net Worth", format_inr(s["net_worth"])),
            ("Monthly Income", format_inr(s["monthly_income"])),
            ("Monthly Surplus", format_inr(s["monthly_surplus"])),
            ("Health Score", f"{s['health_score']} ({s['health_grade']})"),
            ("Retirement Progress", f"{s['retirement_progress_percent']}%"),
            ("Recommended Tax Regime", s["recommended_tax_regime"].upper()),
            ("Goals Tracked", s["total_goals"]),
            ("Active Loans", s["total_loans"]),
            ("FIRE Number", format_inr(plan["fire"]["fire_number"])),
            ("Emergency Fund Status", "Adequate" if plan["emergency_fund"]["adequate"] else "Gap Exists"),
        ]
        apply_header_row(ws, 3, 1, 2)
        ws["A3"], ws["B3"] = "Metric", "Value"
        for i, (k, v) in enumerate(items, start=4):
            ws.cell(row=i, column=1, value=k).border = THIN_BORDER
            ws.cell(row=i, column=2, value=v).border = THIN_BORDER
        autosize(ws)

    def _income(self, ws, plan, profile):
        ws["A1"] = "Income Analysis"
        ws["A1"].font = TITLE_FONT
        headers = ["Source", "Monthly (₹)", "Annual (₹)"]
        apply_header_row(ws, 3, 1, 3)
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        inc = profile.income
        rows = [
            ("Salary", inc.salary_monthly, inc.annual_salary),
            ("Bonus (monthly equiv.)", inc.bonus_annual / 12, inc.bonus_annual),
            ("Business", inc.business_income_annual / 12, inc.business_income_annual),
            ("Rental", inc.rental_income_monthly, inc.annual_rental),
            ("Other", inc.other_income_annual / 12, inc.other_income_annual),
        ]
        for r_idx, (name, m, a) in enumerate(rows, start=4):
            ws.cell(row=r_idx, column=1, value=name).border = THIN_BORDER
            ws.cell(row=r_idx, column=2, value=round(m, 2)).border = THIN_BORDER
            ws.cell(row=r_idx, column=3, value=round(a, 2)).border = THIN_BORDER
        total_row = 4 + len(rows)
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=round(inc.total_monthly_income, 2))
        ws.cell(row=total_row, column=3, value=round(inc.total_annual_income, 2))
        for c in range(1, 4):
            ws.cell(row=total_row, column=c).fill = ACCENT_FILL
            ws.cell(row=total_row, column=c).border = THIN_BORDER
        autosize(ws)

    def _expenses(self, ws, plan, profile):
        ws["A1"] = "Expense Analysis"
        ws["A1"].font = TITLE_FONT
        apply_header_row(ws, 3, 1, 3)
        ws["A3"], ws["B3"], ws["C3"] = "Category", "Monthly (₹)", "Annual (₹)"
        exp = profile.expenses
        rows = [
            ("Living", exp.monthly_living),
            ("Travel", exp.travel_monthly),
            ("Medical", exp.medical_monthly),
            ("Education", exp.education_monthly),
            ("Insurance Premiums", exp.insurance_premium_monthly),
            ("Entertainment", exp.entertainment_monthly),
            ("Miscellaneous", exp.miscellaneous_monthly),
            ("Rent", exp.rent_monthly),
        ]
        for i, (name, m) in enumerate(rows, start=4):
            ws.cell(row=i, column=1, value=name).border = THIN_BORDER
            ws.cell(row=i, column=2, value=m).border = THIN_BORDER
            ws.cell(row=i, column=3, value=m * 12).border = THIN_BORDER
        tr = 4 + len(rows)
        ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=tr, column=2, value=exp.total_monthly)
        ws.cell(row=tr, column=3, value=exp.total_annual)
        for c in range(1, 4):
            ws.cell(row=tr, column=c).fill = ACCENT_FILL
            ws.cell(row=tr, column=c).border = THIN_BORDER
        autosize(ws)

    def _assets(self, ws, plan, profile):
        ws["A1"] = "Assets"
        ws["A1"].font = TITLE_FONT
        apply_header_row(ws, 3, 1, 2)
        ws["A3"], ws["B3"] = "Asset Class", "Amount (₹)"
        a = profile.assets
        items = [
            ("Savings", a.savings),
            ("Cash", a.cash),
            ("Emergency Fund", a.emergency_fund),
            ("Fixed Deposits", a.fd),
            ("PPF", a.ppf),
            ("EPF", a.epf),
            ("NPS", a.nps),
            ("Mutual Funds", a.mutual_funds),
            ("Stocks", a.stocks),
            ("Gold", a.gold),
            ("REITs", a.reit),
            ("InvITs", a.invit),
            ("Real Estate", a.real_estate),
            ("Other", a.other),
        ]
        for i, (name, val) in enumerate(items, start=4):
            ws.cell(row=i, column=1, value=name).border = THIN_BORDER
            ws.cell(row=i, column=2, value=val).border = THIN_BORDER
        tr = 4 + len(items)
        ws.cell(row=tr, column=1, value="TOTAL ASSETS").font = Font(bold=True)
        ws.cell(row=tr, column=2, value=a.total_assets)
        ws.cell(row=tr, column=1).fill = ACCENT_FILL
        ws.cell(row=tr, column=2).fill = ACCENT_FILL
        autosize(ws)
        ws.auto_filter.ref = f"A3:B{tr - 1}"

    def _liabilities(self, ws, plan):
        ws["A1"] = "Liabilities"
        ws["A1"].font = TITLE_FONT
        headers = ["Loan", "Type", "Outstanding (₹)", "Rate %", "EMI (₹)", "Months Left", "Interest Remaining (₹)"]
        apply_header_row(ws, 3, 1, len(headers))
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        for r, loan in enumerate(plan.get("loans", []), start=4):
            vals = [
                loan["name"],
                loan["loan_type"],
                loan["principal_outstanding"],
                round(loan["interest_rate_annual"] * 100, 2),
                loan["emi"],
                loan["tenure_months_remaining"],
                loan["total_interest_without_prepayment"],
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        autosize(ws)

    def _net_worth(self, ws, plan):
        ws["A1"] = "Net Worth Statement"
        ws["A1"].font = TITLE_FONT
        nw = plan["net_worth"]
        apply_header_row(ws, 3, 1, 2)
        ws["A3"], ws["B3"] = "Item", "Amount (₹)"
        rows = [
            ("Total Assets", nw["total_assets"]),
            ("Financial Assets", nw["financial_assets"]),
            ("Real Estate", nw["real_estate"]),
            ("Total Liabilities", nw["total_liabilities"]),
            ("Net Worth", nw["net_worth"]),
            ("Financial Net Worth", nw["financial_net_worth"]),
        ]
        for i, (k, v) in enumerate(rows, start=4):
            ws.cell(row=i, column=1, value=k).border = THIN_BORDER
            cell = ws.cell(row=i, column=2, value=v)
            cell.border = THIN_BORDER
            if k == "Net Worth":
                cell.fill = GOOD_FILL if v >= 0 else BAD_FILL
        # Named range conceptually via cell comment
        ws["B8"].comment = None
        autosize(ws)

        chart = BarChart()
        chart.title = "Net Worth Components"
        data = Reference(ws, min_col=2, min_row=3, max_row=8)
        cats = Reference(ws, min_col=1, min_row=4, max_row=8)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "D3")

    def _cash_flow(self, ws, plan):
        ws["A1"] = "Cash Flow Statement"
        ws["A1"].font = TITLE_FONT
        cf = plan["cash_flow"]
        ratios = plan["ratios"]
        row = 3
        for k, v in [
            ("Monthly Income", cf["monthly_income"]),
            ("Monthly Expenses", cf["monthly_expenses"]),
            ("Monthly EMI", cf["monthly_emi"]),
            ("Monthly SIP", cf["monthly_sip"]),
            ("Monthly Surplus", cf["monthly_surplus"]),
            ("Investable Surplus", cf["investable_surplus"]),
            ("Savings Ratio %", ratios["savings_ratio_percent"]),
            ("DTI %", ratios["debt_to_income_percent"]),
        ]:
            row = write_kv(ws, row, k, v)
        autosize(ws)

    def _emergency(self, ws, plan):
        ws["A1"] = "Emergency Fund Analysis"
        ws["A1"].font = TITLE_FONT
        ef = plan["emergency_fund"]
        row = 3
        for k, v in ef.items():
            if k == "recommendation":
                continue
            row = write_kv(ws, row, k.replace("_", " ").title(), v)
        row = write_kv(ws, row + 1, "Recommendation", ef["recommendation"])
        status_cell = ws.cell(row=3, column=2)
        if ef["adequate"]:
            status_cell.fill = GOOD_FILL
        else:
            status_cell.fill = WARN_FILL
        autosize(ws)

    def _insurance(self, ws, plan):
        ws["A1"] = "Insurance Need Analysis"
        ws["A1"].font = TITLE_FONT
        ins = plan["insurance"]
        write_kv(ws, 3, "Human Life Value", ins["human_life_value"])
        write_kv(ws, 4, "Income Replacement Need", ins["income_replacement_need"])
        headers = ["Cover Type", "Existing", "Recommended", "Gap", "Adequate"]
        apply_header_row(ws, 6, 1, 5)
        for i, h in enumerate(headers, 1):
            ws.cell(row=6, column=i, value=h)
        for r, key in enumerate(["life", "health", "critical_illness", "accident"], start=7):
            d = ins[key]
            vals = [key.replace("_", " ").title(), d["existing_cover"], d["recommended_cover"], d["gap"], "Yes" if d["adequate"] else "No"]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = THIN_BORDER
                if c == 5:
                    cell.fill = GOOD_FILL if d["adequate"] else BAD_FILL
        row = 12
        ws.cell(row=row, column=1, value="Recommendations").font = SECTION_FONT
        for rec in ins["recommendations"]:
            row += 1
            ws.cell(row=row, column=1, value=rec)
        autosize(ws)

    def _portfolio(self, ws, plan):
        ws["A1"] = "Investment Portfolio"
        ws["A1"].font = TITLE_FONT
        headers = ["Name", "Category", "Amount (₹)", "Monthly SIP (₹)", "Expected Return", "Projected 1Y"]
        apply_header_row(ws, 3, 1, 6)
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        for r, inv in enumerate(plan["portfolio"].get("investments", []), start=4):
            vals = [
                inv["name"],
                inv["category"],
                inv["amount"],
                inv["monthly_sip"],
                f"{inv['expected_return']*100:.1f}%",
                inv["projected_1y"],
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        autosize(ws)
        ws.auto_filter.ref = "A3:F100"

    def _allocation(self, ws, plan):
        ws["A1"] = "Asset Allocation"
        ws["A1"].font = TITLE_FONT
        current = plan["portfolio"]["current"].get("asset_class_percent", {})
        target = plan["portfolio"]["target"].get("target_allocation", {})
        apply_header_row(ws, 3, 1, 3)
        ws["A3"], ws["B3"], ws["C3"] = "Asset Class", "Current %", "Target %"
        all_keys = sorted(set(list(current.keys()) + list(target.keys())))
        for i, key in enumerate(all_keys, start=4):
            ws.cell(row=i, column=1, value=key.replace("_", " ").title()).border = THIN_BORDER
            ws.cell(row=i, column=2, value=current.get(key, 0)).border = THIN_BORDER
            tgt = target.get(key, 0)
            ws.cell(row=i, column=3, value=round(tgt * 100, 2) if tgt <= 1 else tgt).border = THIN_BORDER

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(all_keys))
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(all_keys))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Current Allocation"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        ws.add_chart(pie, "E3")

        ws.cell(row=4 + len(all_keys) + 2, column=1, value="Rebalancing Needs").font = SECTION_FONT
        for i, item in enumerate(plan["portfolio"].get("rebalancing", []), start=4 + len(all_keys) + 3):
            ws.cell(row=i, column=1, value=f"{item['asset_class']}: {item['action']} by {abs(item['difference_percent'])}%")
        autosize(ws)

    def _filtered_investments(self, ws, plan, profile, title, categories):
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        apply_header_row(ws, 3, 1, 4)
        for i, h in enumerate(["Name", "Amount (₹)", "SIP (₹)", "Return"], 1):
            ws.cell(row=3, column=i, value=h)
        row = 4
        for inv in profile.investments:
            if inv.category.lower() in categories or title.lower().split()[0].lower() in inv.category.lower():
                for c, v in enumerate([inv.name, inv.amount, inv.monthly_sip, inv.expected_return], 1):
                    ws.cell(row=row, column=c, value=v).border = THIN_BORDER
                row += 1
        if row == 4:
            # Fall back to asset totals
            assets = profile.assets
            mapping = {
                "Mutual Funds": assets.mutual_funds,
                "Stocks": assets.stocks,
                "Gold": assets.gold,
            }
            ws.cell(row=4, column=1, value=title).border = THIN_BORDER
            ws.cell(row=4, column=2, value=mapping.get(title, 0)).border = THIN_BORDER
        autosize(ws)

    def _mutual_funds(self, ws, plan, profile):
        self._filtered_investments(ws, plan, profile, "Mutual Funds", {"mutual_funds", "equity", "hybrid", "debt"})

    def _stocks(self, ws, plan, profile):
        self._filtered_investments(ws, plan, profile, "Stocks", {"stocks", "equity"})

    def _gold(self, ws, plan, profile):
        self._filtered_investments(ws, plan, profile, "Gold", {"gold"})

    def _loans(self, ws, plan):
        ws["A1"] = "Loan Summary & Prepayment Analysis"
        ws["A1"].font = TITLE_FONT
        headers = [
            "Name", "Type", "Outstanding", "Rate %", "EMI", "Interest w/o Prep",
            "Interest w/ Prep", "Interest Saved", "Months Saved",
        ]
        apply_header_row(ws, 3, 1, len(headers))
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        for r, loan in enumerate(plan.get("loans", []), start=4):
            vals = [
                loan["name"],
                loan["loan_type"],
                loan["principal_outstanding"],
                round(loan["interest_rate_annual"] * 100, 2),
                loan["emi"],
                loan["total_interest_without_prepayment"],
                loan["total_interest_with_prepayment"],
                loan["interest_saved"],
                loan["months_saved"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = THIN_BORDER
                if c == 8 and isinstance(v, (int, float)) and v > 0:
                    cell.fill = GOOD_FILL
        autosize(ws)

    def _loan_schedule(self, ws, plan):
        ws["A1"] = "Amortization Schedule"
        ws["A1"].font = TITLE_FONT
        headers = ["Loan", "Month", "Opening", "EMI", "Interest", "Principal", "Prepayment", "Closing"]
        apply_header_row(ws, 3, 1, len(headers))
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        row = 4
        # Use full amortization from loans_full if present
        loans = plan.get("loans_full") or plan.get("loans", [])
        for loan in loans:
            schedule = loan.get("amortization", loan.get("amortization_preview", []))
            for entry in schedule[:360]:
                vals = [
                    loan["name"],
                    entry.get("month"),
                    entry.get("opening_balance"),
                    entry.get("emi"),
                    entry.get("interest"),
                    entry.get("principal"),
                    entry.get("prepayment"),
                    entry.get("closing_balance"),
                ]
                for c, v in enumerate(vals, 1):
                    ws.cell(row=row, column=c, value=v).border = THIN_BORDER
                row += 1
        ws.auto_filter.ref = f"A3:H{max(row - 1, 3)}"
        autosize(ws)

    def _goals(self, ws, plan):
        ws["A1"] = "Goal Planning"
        ws["A1"].font = TITLE_FONT
        headers = [
            "Goal", "Type", "Priority", "Current Cost", "Future Cost", "Years",
            "Monthly SIP", "Lumpsum", "Progress %", "Gap",
        ]
        apply_header_row(ws, 3, 1, len(headers))
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        for r, g in enumerate(plan.get("goals", []), start=4):
            vals = [
                g["name"], g["goal_type"], g["priority"], g["current_cost"],
                g["future_cost"], g["years_to_goal"], g["monthly_sip_required"],
                g["lumpsum_required"], g["progress_percent"], g["funding_gap"],
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        # Conditional formatting on progress
        if plan.get("goals"):
            ws.conditional_formatting.add(
                f"I4:I{3 + len(plan['goals'])}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="FEE2E2",
                    mid_type="num", mid_value=50, mid_color="FEF3C7",
                    end_type="num", end_value=100, end_color="D1FAE5",
                ),
            )
        # Priority dropdown
        dv = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"C4:C{3 + max(len(plan.get('goals', [])), 1)}")
        autosize(ws)
        ws.auto_filter.ref = "A3:J100"

    def _retirement(self, ws, plan):
        ws["A1"] = "Retirement Planning"
        ws["A1"].font = TITLE_FONT
        ret = plan["retirement"]
        row = 3
        keys = [
            "years_to_retirement", "years_in_retirement", "required_corpus",
            "projected_corpus", "shortfall", "surplus",
            "additional_monthly_sip_required", "progress_percent",
            "monthly_retirement_income_needed", "safe_withdrawal_rate",
            "corpus_sustainability_years", "annual_expense_at_retirement",
            "healthcare_cost_at_retirement", "passive_income_at_retirement",
        ]
        for k in keys:
            row = write_kv(ws, row, k.replace("_", " ").title(), ret.get(k))
        row += 2
        ws.cell(row=row, column=1, value="Bucket Strategy").font = SECTION_FONT
        buckets = ret.get("bucket_strategy", {})
        row += 1
        for bname, bdata in buckets.items():
            ws.cell(row=row, column=1, value=bname.replace("_", " ").title())
            ws.cell(row=row, column=2, value=bdata.get("amount"))
            ws.cell(row=row, column=3, value=bdata.get("purpose"))
            row += 1
        row += 1
        mc = ret.get("monte_carlo", {})
        ws.cell(row=row, column=1, value="Monte Carlo Simulation").font = SECTION_FONT
        row += 1
        for k in ("success_rate_percent", "median_ending_corpus", "percentile_10", "percentile_90", "interpretation"):
            row = write_kv(ws, row, k.replace("_", " ").title(), mc.get(k))
        autosize(ws)

    def _tax(self, ws, plan):
        ws["A1"] = "Tax Planning — Old vs New Regime"
        ws["A1"].font = TITLE_FONT
        tax = plan["tax"]
        write_kv(ws, 3, "Recommended Regime", tax.get("recommended_regime", "").upper())
        write_kv(ws, 4, "Tax Savings", tax.get("tax_savings_by_switching"))
        write_kv(ws, 5, "Recommendation", tax.get("recommendation_text"))

        apply_header_row(ws, 7, 1, 3)
        ws["A7"], ws["B7"], ws["C7"] = "Metric", "Old Regime", "New Regime"
        old, new = tax["old_regime"], tax["new_regime"]
        metrics = [
            "gross_income", "standard_deduction", "total_deductions",
            "taxable_income", "tax_before_rebate", "rebate_87a",
            "ltcg_tax", "stcg_tax", "cess", "total_tax", "effective_tax_rate",
        ]
        for i, m in enumerate(metrics, start=8):
            ws.cell(row=i, column=1, value=m.replace("_", " ").title()).border = THIN_BORDER
            ws.cell(row=i, column=2, value=old.get(m)).border = THIN_BORDER
            ws.cell(row=i, column=3, value=new.get(m)).border = THIN_BORDER
        # Highlight lower tax
        if old["total_tax"] < new["total_tax"]:
            ws.cell(row=8 + metrics.index("total_tax"), column=2).fill = GOOD_FILL
        else:
            ws.cell(row=8 + metrics.index("total_tax"), column=3).fill = GOOD_FILL
        autosize(ws)

    def _goal_type_sheet(self, ws, plan, title, goal_types):
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        goals = [g for g in plan.get("goals", []) if g["goal_type"] in goal_types]
        headers = ["Goal", "Current Cost", "Future Cost", "Monthly SIP", "Progress %"]
        apply_header_row(ws, 3, 1, 5)
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        if not goals:
            ws["A4"] = "No goals of this type configured. Add goals via the client input form."
        for r, g in enumerate(goals, start=4):
            for c, v in enumerate(
                [g["name"], g["current_cost"], g["future_cost"], g["monthly_sip_required"], g["progress_percent"]],
                1,
            ):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        autosize(ws)

    def _education(self, ws, plan):
        self._goal_type_sheet(ws, plan, "Education Planning", {"children_education"})

    def _marriage(self, ws, plan):
        self._goal_type_sheet(ws, plan, "Marriage Planning", {"children_marriage"})

    def _travel(self, ws, plan):
        self._goal_type_sheet(ws, plan, "Travel Planning", {"vacation"})

    def _estate(self, ws, plan):
        ws["A1"] = "Estate Planning Checklist"
        ws["A1"].font = TITLE_FONT
        headers = ["Item", "Priority", "Description", "Status"]
        apply_header_row(ws, 3, 1, 4)
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        for r, item in enumerate(plan.get("estate_checklist", []), start=4):
            for c, v in enumerate(
                [item["item"], item["priority"], item["description"], item["status"]], 1
            ):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        dv = DataValidation(type="list", formula1='"pending,in_progress,complete,not_applicable"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"D4:D{3 + max(len(plan.get('estate_checklist', [])), 1)}")
        autosize(ws)

    def _recommendations(self, ws, plan):
        ws["A1"] = "Recommendations & Action Plan"
        ws["A1"].font = TITLE_FONT
        ws["A3"] = "Priority Actions"
        ws["A3"].font = SECTION_FONT
        apply_header_row(ws, 4, 1, 3)
        ws["A4"], ws["B4"], ws["C4"] = "Priority", "Area", "Action"
        for r, action in enumerate(plan["recommendations"].get("priority_actions", []), start=5):
            for c, v in enumerate([action["priority"], action["area"], action["action"]], 1):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER

        start = 5 + len(plan["recommendations"].get("priority_actions", [])) + 2
        ws.cell(row=start, column=1, value="Investment Allocation Recommendations").font = SECTION_FONT
        apply_header_row(ws, start + 1, 1, 5)
        for i, h in enumerate(["Category", "Allocation %", "Monthly ₹", "Annual ₹", "Description"], 1):
            ws.cell(row=start + 1, column=i, value=h)
        for r, rec in enumerate(
            plan["recommendations"].get("investment", {}).get("recommendations", []),
            start=start + 2,
        ):
            for c, v in enumerate(
                [rec["name"], rec["allocation_percent"], rec["monthly_amount"], rec["annual_amount"], rec["description"]],
                1,
            ):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        autosize(ws)

    def _risk(self, ws, plan):
        ws["A1"] = "Risk Analysis"
        ws["A1"].font = TITLE_FONT
        risk = plan["risk"]
        write_kv(ws, 3, "Profile", risk.get("profile"))
        write_kv(ws, 4, "Personality", risk.get("personality"))
        write_kv(ws, 5, "Score", risk.get("total_score"))
        write_kv(ws, 6, "Description", risk.get("description"))
        apply_header_row(ws, 8, 1, 3)
        ws["A8"], ws["B8"], ws["C8"] = "Question", "Selected", "Score"
        for r, d in enumerate(risk.get("answers_detail", []), start=9):
            ws.cell(row=r, column=1, value=d.get("question")).border = THIN_BORDER
            ws.cell(row=r, column=2, value=d.get("selected_option")).border = THIN_BORDER
            ws.cell(row=r, column=3, value=d.get("score")).border = THIN_BORDER
        autosize(ws)

    def _scenarios(self, ws, plan):
        ws["A1"] = "Scenario Analysis"
        ws["A1"].font = TITLE_FONT
        apply_header_row(ws, 3, 1, 5)
        for i, h in enumerate(["Scenario", "Monthly Surplus", "Required Corpus", "Projected Corpus", "Shortfall"], 1):
            ws.cell(row=3, column=i, value=h)
        for r, (key, data) in enumerate(plan.get("scenarios", {}).items(), start=4):
            vals = [
                data.get("label", key),
                data.get("monthly_surplus"),
                data.get("required_corpus"),
                data.get("projected_corpus"),
                data.get("shortfall"),
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        chart = BarChart()
        chart.title = "Projected vs Required Corpus"
        data = Reference(ws, min_col=3, min_row=3, max_col=4, max_row=6)
        cats = Reference(ws, min_col=1, min_row=4, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A10")
        autosize(ws)

    def _inflation(self, ws, plan):
        ws["A1"] = "Inflation Analysis"
        ws["A1"].font = TITLE_FONT
        infl = plan["assumptions_used"].get("inflation", {})
        row = 3
        for k, v in infl.items():
            row = write_kv(ws, row, k.replace("_", " ").title(), f"{v*100:.1f}%")
        row += 2
        ws.cell(row=row, column=1, value="Goal Future Costs (Inflation Adjusted)").font = SECTION_FONT
        row += 1
        apply_header_row(ws, row, 1, 4)
        for i, h in enumerate(["Goal", "Current", "Inflation", "Future"], 1):
            ws.cell(row=row, column=i, value=h)
        row += 1
        for g in plan.get("goals", []):
            for c, v in enumerate(
                [g["name"], g["current_cost"], f"{g['inflation_rate']*100:.1f}%", g["future_cost"]], 1
            ):
                ws.cell(row=row, column=c, value=v).border = THIN_BORDER
            row += 1
        autosize(ws)

    def _dashboard(self, ws, plan):
        ws["A1"] = "Executive Dashboard"
        ws["A1"].font = TITLE_FONT
        s = plan["summary"]
        hs = plan["health_score"]
        metrics = [
            ("Net Worth", format_inr(s["net_worth"])),
            ("Monthly Surplus", format_inr(s["monthly_surplus"])),
            ("Health Score", f"{hs['score']} / 100 ({hs['grade']})"),
            ("Retirement Progress", f"{s['retirement_progress_percent']}%"),
            ("Savings Ratio", f"{plan['ratios']['savings_ratio_percent']}%"),
            ("DTI Ratio", f"{plan['ratios']['debt_to_income_percent']}%"),
            ("FIRE Progress", f"{plan['fire']['current_progress_percent']}%"),
            ("Tax Regime", s["recommended_tax_regime"].upper()),
        ]
        apply_header_row(ws, 3, 1, 2)
        ws["A3"], ws["B3"] = "KPI", "Value"
        for i, (k, v) in enumerate(metrics, start=4):
            ws.cell(row=i, column=1, value=k).border = THIN_BORDER
            ws.cell(row=i, column=2, value=v).border = THIN_BORDER
            ws.cell(row=i, column=1).fill = LIGHT_FILL

        ws["D3"] = "Top Improvements"
        ws["D3"].font = SECTION_FONT
        for i, tip in enumerate(hs.get("top_improvements", []), start=4):
            ws.cell(row=i, column=4, value=f"{i - 3}. {tip}")
        autosize(ws)

    def _charts_sheet(self, ws, plan):
        ws["A1"] = "Charts Data"
        ws["A1"].font = TITLE_FONT
        ws["A3"] = "Asset Class"
        ws["B3"] = "Amount"
        apply_header_row(ws, 3, 1, 2)
        ac = plan["portfolio"]["current"].get("asset_class", {})
        for i, (k, v) in enumerate(ac.items(), start=4):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)
        pie = PieChart()
        pie.title = "Portfolio Mix"
        labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(ac))
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(ac))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, "D3")
        autosize(ws)

    def _assumptions(self, ws, plan):
        ws["A1"] = "Assumptions Used"
        ws["A1"].font = TITLE_FONT
        row = 3
        for section, data in plan.get("assumptions_used", {}).items():
            ws.cell(row=row, column=1, value=str(section).upper()).font = SECTION_FONT
            row += 1
            if isinstance(data, dict):
                for k, v in data.items():
                    row = write_kv(ws, row, str(k), v)
            else:
                row = write_kv(ws, row, str(section), data)
            row += 1
        autosize(ws)

    def _glossary(self, ws):
        ws["A1"] = "Glossary"
        ws["A1"].font = TITLE_FONT
        terms = [
            ("SIP", "Systematic Investment Plan — fixed periodic investment into mutual funds."),
            ("XIRR", "Extended Internal Rate of Return for irregular cash flows."),
            ("CAGR", "Compound Annual Growth Rate."),
            ("HLV", "Human Life Value — present value of future earnings for dependents."),
            ("SWR", "Safe Withdrawal Rate — sustainable annual withdrawal percentage."),
            ("FIRE", "Financial Independence, Retire Early — corpus = expenses / SWR."),
            ("DTI", "Debt-to-Income ratio — EMI / monthly income."),
            ("PPF", "Public Provident Fund — EEE tax-exempt long-term debt instrument."),
            ("NPS", "National Pension System — retirement account with tax benefits."),
            ("EPF", "Employees' Provident Fund — mandatory salaried retirement savings."),
            ("LTCG", "Long-Term Capital Gains."),
            ("STCG", "Short-Term Capital Gains."),
            ("REIT", "Real Estate Investment Trust."),
            ("InvIT", "Infrastructure Investment Trust."),
            ("Monte Carlo", "Stochastic simulation of portfolio outcomes under return uncertainty."),
        ]
        apply_header_row(ws, 3, 1, 2)
        ws["A3"], ws["B3"] = "Term", "Definition"
        for i, (t, d) in enumerate(terms, start=4):
            ws.cell(row=i, column=1, value=t).border = THIN_BORDER
            ws.cell(row=i, column=2, value=d).border = THIN_BORDER
        autosize(ws, max_width=80)
