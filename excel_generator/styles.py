"""Professional Excel styling constants."""

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

# Brand palette — deep teal / navy wealth-management look
NAVY = "0B3D5C"
TEAL = "1F6F8B"
ACCENT = "99C24D"
WARM = "F18F01"
DANGER = "C73E1D"
LIGHT_BG = "F4F7FA"
HEADER_BG = "0B3D5C"
ALT_ROW = "E8F1F5"
WHITE = "FFFFFF"
BLACK = "1A1A1A"
GRAY = "64748B"

thin = Side(style="thin", color="CBD5E1")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=BLACK)
VALUE_FONT = Font(name="Calibri", size=10, color=BLACK)
MONEY_FONT = Font(name="Calibri", size=10, color=BLACK)

HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
SECTION_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=ALT_ROW)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT_BG)
ACCENT_FILL = PatternFill("solid", fgColor=ACCENT)
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
GOOD_FILL = PatternFill("solid", fgColor="D1FAE5")
BAD_FILL = PatternFill("solid", fgColor="FEE2E2")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def inr_format(value: float) -> str:
    return f"₹{value:,.0f}"


def apply_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autosize(ws, min_width: int = 12, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def write_kv(ws, row: int, label: str, value, label_col: int = 1, value_col: int = 2) -> int:
    ws.cell(row=row, column=label_col, value=label).font = LABEL_FONT
    ws.cell(row=row, column=label_col).fill = LIGHT_FILL
    ws.cell(row=row, column=label_col).border = THIN_BORDER
    cell = ws.cell(row=row, column=value_col, value=value)
    cell.font = VALUE_FONT
    cell.border = THIN_BORDER
    cell.alignment = LEFT
    return row + 1
